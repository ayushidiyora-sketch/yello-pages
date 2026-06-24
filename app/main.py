import asyncio
import uuid
from contextlib import asynccontextmanager
from datetime import datetime

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, HTMLResponse, Response

from .db import (jobs, businesses, products, reviews, ebay_products, gresults, bbbresults,
                 g2reviews, bbbreviews, gjobs, gcompanies, gdreviews, walmart_products, walmart_reviews,
                 youtube_channels, airbnb_reviews, expedia_results, trustpilot_results,
                 hotels_results, hotels_reviews, trustpilot_search_results, homedepot_results,
                 trustpilot_reviews, monitors, expedia_reviews, airbnb_search_results,
                 gmaps_results, gmaps_domain_results, gmaps_reviews, gnews_results,
                 gmaps_contrib_reviews, gmaps_photos, gimages_results, gmaps_traffic,
                 gmaps_directory, gvideos_results, gevents_results, gcareers_results,
                 gtrends_results, linkedin_companies_results, linkedin_posts_results,
                 gsjobs_results, gshop_results, gplay_results,
                 gsreviews_results, linkedin_profiles, gflights_results, gmaps_autocomplete,
                 gsearch_autocomplete,
                 ensure_indexes)
from .models import (ScrapeRequest, AmazonScrapeRequest, AmazonReviewsRequest,
                     EbayScrapeRequest, GSearchRequest, BBBRequest, G2Request, BBBReviewsRequest,
                     GlassdoorJobsRequest, GlassdoorCompaniesRequest, GlassdoorReviewsRequest,
                     WalmartProductsRequest,
                     WalmartReviewsRequest, YouTubeChannelsRequest, AirbnbReviewsRequest,
                     ExpediaRequest, TrustpilotRequest, HotelsRequest, HotelsReviewsRequest,
                     TrustpilotSearchRequest, HomeDepotRequest, TrustpilotReviewsRequest,
                     TrustpilotMonitorRequest, ExpediaReviewsRequest, AirbnbSearchRequest,
                     GoogleMapsRequest, GoogleMapsDomainsRequest, GMapsReviewsRequest,
                     GMapsMonitorRequest, GNewsRequest, GMapsContribRequest, GMapsPhotosRequest,
                     GImagesRequest, GMapsTrafficRequest, GMapsDirectoryRequest, GVideosRequest,
                     GEventsRequest, GCareersRequest, GTrendsRequest, LinkedInCompaniesRequest,
                     LinkedInPostsRequest,
                     GSJobsRequest, GShopRequest, GShopReviewsRequest, GPlayRequest,
                     GPlayMonitorRequest, LinkedInProfilesRequest, GFlightsRequest,
                     GMapsAutocompleteRequest, GSearchAutocompleteRequest)
from .scraper import run_scrape, request_stop, apply_view, REGIONS, SUPPORTED_REGIONS
from . import (yp_us, amazon, amazon_reviews, ebay, gsearch, bbb, bbb_reviews, g2,
               glassdoor_jobs, glassdoor_companies, glassdoor_reviews, walmart, walmart_reviews as walmart_rv,
               youtube_channels as yt_channels, airbnb_reviews as airbnb_rv,
               expedia, trustpilot, hotels, hotels_reviews as hotels_reviews_mod,
               trustpilot_search, homedepot, trustpilot_reviews as trustpilot_reviews_mod,
               monitor, storage, expedia_reviews as expedia_reviews_mod, airbnb, gmaps,
               gmaps_reviews as gmaps_reviews_mod, gnews, gmaps_contrib,
               gmaps_photos as gmaps_photos_mod, gimages, gmaps_traffic as gmaps_traffic_mod,
               gmaps_directory as gmaps_directory_mod, gvideos, gevents, gcareers, gtrends,
               linkedin_companies, linkedin_posts,
               gsjobs, gshop, gsreviews, gplay,
               linkedin_profiles as linkedin_profiles_mod, gflights,
               gmaps_autocomplete as gmaps_autocomplete_mod)


# ---------------- auto-save each finished job to data/<service>/<job>/results.xlsx ----------------
# A thin wrapper around every background scrape: run it as usual (Mongo stays the live store
# the UI polls), then — done, stopped, or error — export whatever landed in Mongo to a per-job
# folder with an Excel + JSON file. Scraper modules are untouched.

async def _amazon_rows(job_id):
    rows = [amazon.to_export(d) async for d in products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    fixed = set(amazon.EXPORT_COLUMNS)
    extras = sorted({k for r in rows for k in r if k not in fixed})
    header = [c for c in amazon.EXPORT_COLUMNS if c not in ("position", "query")] \
        + extras + ["position", "query"]
    return rows, header


async def _reviews_rows(job_id):
    rows = [amazon_reviews.to_export(d) async for d in reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, amazon_reviews.REVIEW_COLUMNS


async def _ebay_rows(job_id):
    rows = [ebay.to_export(d) async for d in ebay_products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, ebay.EBAY_COLUMNS


async def _yp_rows(job_id):
    rows = [d async for d in businesses.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, None


async def _gsearch_rows(job_id):
    rows = [d async for d in gresults.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, None


async def _gnews_rows(job_id):
    rows = [d async for d in gnews_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gnews.GNEWS_COLUMNS


async def _gimages_rows(job_id):
    rows = [d async for d in gimages_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gimages.GIMG_COLUMNS


async def _gvideos_rows(job_id):
    rows = [d async for d in gvideos_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gvideos.GVID_COLUMNS


async def _gevents_rows(job_id):
    rows = [d async for d in gevents_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gevents.GEVENTS_COLUMNS


async def _gcareers_rows(job_id):
    rows = [d async for d in gcareers_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gcareers.GCAREERS_COLUMNS


async def _gtrends_rows(job_id):
    rows = [d async for d in gtrends_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gtrends.GTRENDS_COLUMNS


async def _linkedin_companies_rows(job_id):
    rows = [d async for d in linkedin_companies_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, linkedin_companies.LINKEDIN_COMPANY_COLUMNS


async def _linkedin_posts_rows(job_id):
    rows = [d async for d in linkedin_posts_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, linkedin_posts.LINKEDIN_POSTS_COLUMNS


async def _gsjobs_rows(job_id):
    rows = [d async for d in gsjobs_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gsjobs.GSJ_COLUMNS


async def _gshop_rows(job_id):
    rows = [d async for d in gshop_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gshop.GSH_COLUMNS


async def _gsr_rows(job_id):
    rows = [d async for d in gsreviews_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gsreviews.GSR_COLUMNS


async def _lip_rows(job_id):
    rows = [d async for d in linkedin_profiles.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, linkedin_profiles_mod.LIP_COLUMNS


async def _gfl_rows(job_id):
    rows = [d async for d in gflights_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gflights.GFL_COLUMNS


async def _gma_rows(job_id):
    rows = [d async for d in gmaps_autocomplete.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gmaps_autocomplete_mod.GMA_COLUMNS


async def _gsa_rows(job_id):
    rows = [d async for d in gsearch_autocomplete.find(
        {"job_id": job_id}, {"_id": 0, "job_id": 0, "coordinates": 0})]
    return rows, ["query", "suggestion", "position"]


async def _gplay_rows(job_id):
    rows = [d async for d in gplay_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gplay.GPL_COLUMNS


async def _gmaps_traffic_rows(job_id):
    rows = [d async for d in gmaps_traffic.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gmaps_traffic_mod.GMT_COLUMNS


async def _gmaps_directory_rows(job_id):
    rows = [d async for d in gmaps_directory.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gmaps_directory_mod.GMD_COLUMNS


async def _gmaps_contrib_rows(job_id):
    rows = [d async for d in gmaps_contrib_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gmaps_contrib.GMCR_COLUMNS


async def _gmaps_photos_rows(job_id):
    rows = [d async for d in gmaps_photos.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, gmaps_photos_mod.GMP_COLUMNS


async def _bbb_rows(job_id):
    rows = [d async for d in bbbresults.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, None


async def _bbbreviews_rows(job_id):
    rows = [d async for d in bbbreviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows, None


async def _g2_rows(job_id):
    rows = [g2.to_export(d) async for d in g2reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, g2.G2_COLUMNS


async def _gjobs_rows(job_id):
    rows = [glassdoor_jobs.to_export(d) async for d in gjobs.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, glassdoor_jobs.GLASSDOOR_JOB_COLUMNS


async def _gco_rows(job_id):
    rows = [d async for d in gcompanies.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, glassdoor_companies.GLASSDOOR_COMPANY_COLUMNS


async def _gdreviews_rows(job_id):
    rows = [glassdoor_reviews.to_export(d) async for d in gdreviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, glassdoor_reviews.GLASSDOOR_REVIEW_COLUMNS


async def _walmart_rows(job_id):
    rows = [walmart.to_export(d) async for d in walmart_products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, walmart.WALMART_PRODUCT_COLUMNS


async def _walmartrv_rows(job_id):
    rows = [walmart_rv.to_export(d) async for d in walmart_reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, walmart_rv.WALMART_REVIEW_COLUMNS


async def _ytchannels_rows(job_id):
    rows = [yt_channels.to_export(d) async for d in youtube_channels.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, yt_channels.YOUTUBE_CHANNEL_COLUMNS


async def _airbnb_rows(job_id):
    rows = [airbnb_rv.to_export(d) async for d in airbnb_reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows, airbnb_rv.AIRBNB_REVIEW_COLUMNS


async def _run_and_archive(coro, service, job_id, fetch):
    try:
        await coro
    finally:
        try:
            rows, header = await fetch(job_id)
            await storage.archive(service, job_id, rows, header)
        except Exception:
            pass  # archiving must never break the job


@asynccontextmanager
async def lifespan(app: FastAPI):
    await ensure_indexes()
    loop_task = asyncio.create_task(monitor.monitor_loop())  # recurring Trustpilot monitors
    yield
    loop_task.cancel()


app = FastAPI(title="YellowPages US Scraper", lifespan=lifespan)


@app.get("/", response_class=HTMLResponse)
async def index():
    with open("static/index.html", encoding="utf-8") as f:
        return f.read()


@app.get("/bbb-seal.svg")
async def bbb_seal():
    """BBB Accredited Business seal, served locally (bbb.org blocks hot-linking)."""
    return FileResponse("static/bbb-seal.svg", media_type="image/svg+xml")


@app.get("/api/services")
async def services():
    """The Live Scraper service catalog (from the source sheet) for the sidebar."""
    import json
    with open("static/services.json", encoding="utf-8") as f:
        return json.load(f)


@app.get("/api/regions")
async def regions():
    """Region options for the dropdown. `supported` = parser implemented today."""
    labels = {
        "us": "US (yellowpages.com)",
        "au": "AU (yellowpages.com.au)",
        "ca": "CA (yellowpages.ca)",
        "be": "BE (goldenpages.be)",
        "fr": "FR (pagesjaunes.fr)",
    }
    return [
        {"code": code, "label": labels.get(code, code), "supported": code in SUPPORTED_REGIONS}
        for code in REGIONS
    ]


@app.get("/api/filters")
async def filters(search: str, location: str, region: str = "us"):
    """YP's own filter options (Category / Features / Neighborhoods) for this search,
    used to populate the All Filters modal. Live fetch — may take a few seconds."""
    if region not in SUPPORTED_REGIONS:
        raise HTTPException(400, "Filters are only available for the US region.")
    return await yp_us.get_filters(search, location)


@app.post("/api/scrape")
async def start_scrape(req: ScrapeRequest):
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id,
        "search": req.search,
        "location": req.location,
        "region": req.region,
        "limit": req.limit,
        "sort": req.sort,
        "categories": req.categories,
        "status": "running",
        "total_scraped": 0,
        "started_at": datetime.utcnow(),
        "finished_at": None,
    })
    # launch background scraping (returns immediately)
    asyncio.create_task(_run_and_archive(
        run_scrape(job_id, req.search, req.location, req.region, req.limit),
        "yellowpages", job_id, _yp_rows))
    return {"job_id": job_id}


@app.post("/api/amazon/scrape")
async def start_amazon_scrape(req: AmazonScrapeRequest):
    queries = [q.strip() for q in (req.queries or []) if q and q.strip()]
    if not queries:
        raise HTTPException(400, "Provide at least one ASIN, product URL, or search.")
    job_id = uuid.uuid4().hex
    # Expected total for the progress bar: a direct ASIN/product URL yields 1 product;
    # a search query yields up to `limit`. (So 3 product URLs => 3, not 3 x limit.)
    expected = 0
    for q in queries:
        spec = amazon.classify(q, req.domain)
        if not spec:
            continue
        expected += 1 if spec[0] in ("asin", "product", "product_url") else max(1, req.limit)
    await jobs.insert_one({
        "job_id": job_id,
        "kind": "amazon",
        "domain": req.domain,
        "postcode": req.postcode,
        "language": req.language,
        "currency": req.currency,
        "limit": req.limit,
        "queries": queries,
        "status": "running",
        "total_scraped": 0,
        "total_available": expected or len(queries),
        "started_at": datetime.utcnow(),
        "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        amazon.run_amazon_scrape(
            job_id, queries, req.domain, req.postcode, req.language, req.currency, req.limit),
        "amazon_products", job_id, _amazon_rows))
    return {"job_id": job_id}


@app.post("/api/amazon/parse-file")
async def amazon_parse_file(file: UploadFile = File(...)):
    """Extract Amazon product URLs / ASINs from an uploaded CSV/XLSX/TXT/Parquet file
    so the UI can drop them into the Product ASINs/URLs box."""
    data = await file.read()
    try:
        lines = amazon.parse_upload(file.filename or "", data)
    except Exception as e:
        raise HTTPException(400, f"Could not parse '{file.filename}': {e}")
    return {"lines": lines, "count": len(lines)}


@app.get("/api/amazon/results/{job_id}")
async def amazon_results(job_id: str, limit: int = 1000):
    rows = [amazon.to_export(d) async for d in products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows[:limit]


@app.get("/api/amazon/export-excel/{job_id}")
async def amazon_export_excel(job_id: str):
    """Download all scraped products as an .xlsx with the full 95-column schema."""
    import io
    import openpyxl

    rows = [amazon.to_export(d) async for d in products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)

    # header = fixed 95 columns + any dynamic details_/overview_ columns the products have
    fixed = set(amazon.EXPORT_COLUMNS)
    extras = sorted({k for r in rows for k in r if k not in fixed})
    header = [c for c in amazon.EXPORT_COLUMNS if c not in ("position", "query")] \
        + extras + ["position", "query"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(header)
    for r in rows:
        ws.append([_xlsx_cell(r.get(c, "")) for c in header])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="amazon_products_{job_id[:8]}.xlsx"'},
    )


def _xlsx_cell(v):
    """Excel-safe cell value: scalars pass through, anything else becomes a string."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    return str(v)


# ---------------- Amazon Reviews ----------------

@app.post("/api/amazon-reviews/scrape")
async def start_amazon_reviews(req: AmazonReviewsRequest):
    queries = [q.strip() for q in (req.queries or []) if q and q.strip()]
    if not queries:
        raise HTTPException(400, "Provide at least one ASIN or product URL.")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id,
        "kind": "amazon-reviews",
        "domain": req.domain,
        "limit": req.limit,
        "queries": queries,
        "status": "running",
        "total_scraped": 0,
        "total_available": len(queries) * max(1, req.limit),
        "started_at": datetime.utcnow(),
        "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        amazon_reviews.run_reviews_scrape(job_id, queries, req.domain, req.limit),
        "amazon_reviews", job_id, _reviews_rows))
    return {"job_id": job_id}


@app.get("/api/amazon-reviews/results/{job_id}")
async def amazon_reviews_results(job_id: str, limit: int = 2000):
    rows = [amazon_reviews.to_export(d) async for d in reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows[:limit]


@app.get("/api/amazon-reviews/export-excel/{job_id}")
async def amazon_reviews_export_excel(job_id: str):
    """Download all scraped reviews as an .xlsx (one row per review)."""
    import io
    import openpyxl

    rows = [amazon_reviews.to_export(d) async for d in reviews.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reviews"
    ws.append(amazon_reviews.REVIEW_COLUMNS)
    for r in rows:
        ws.append([_xlsx_cell(r.get(c, "")) for c in amazon_reviews.REVIEW_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="amazon_reviews_{job_id[:8]}.xlsx"'},
    )


# ---------------- eBay Products ----------------

@app.post("/api/ebay/scrape")
async def start_ebay_scrape(req: EbayScrapeRequest):
    queries = [q.strip() for q in (req.queries or []) if q and q.strip()]
    if not queries:
        raise HTTPException(400, "Provide at least one eBay item id, URL, or search.")
    job_id = uuid.uuid4().hex
    expected = 0
    for q in queries:
        spec = ebay.classify(q, ebay.domain_for(req.country))
        if not spec:
            continue
        expected += 1 if spec[0] == "item" else max(1, req.limit)
    await jobs.insert_one({
        "job_id": job_id, "kind": "ebay", "country": req.country, "postcode": req.postcode,
        "limit": req.limit, "queries": queries, "status": "running", "total_scraped": 0,
        "total_available": expected or len(queries),
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        ebay.run_ebay_scrape(job_id, queries, req.country, req.postcode, req.limit),
        "ebay_products", job_id, _ebay_rows))
    return {"job_id": job_id}


@app.get("/api/ebay/results/{job_id}")
async def ebay_results(job_id: str, limit: int = 1000):
    rows = [ebay.to_export(d) async for d in ebay_products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    return rows[:limit]


@app.get("/api/ebay/export-excel/{job_id}")
async def ebay_export_excel(job_id: str):
    """Download all scraped eBay products as an .xlsx."""
    import io
    import openpyxl

    rows = [ebay.to_export(d) async for d in ebay_products.find({"job_id": job_id}, {"_id": 0})]
    rows.sort(key=lambda r: r.get("position") or 0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "eBay Products"
    ws.append(ebay.EBAY_COLUMNS)
    for r in rows:
        ws.append([_xlsx_cell(r.get(c, "")) for c in ebay.EBAY_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="ebay_products_{job_id[:8]}.xlsx"'},
    )


@app.post("/api/gsearch")
async def gsearch_start(req: GSearchRequest):
    """Google Search Scraper (powered by DuckDuckGo via the proxy pool — no real IP)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gsearch", "queries": queries, "limit": req.limit,
        "date_range": req.date_range, "region": req.region, "language": req.language,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gsearch.run_job(job_id, queries, req.limit, req.date_range or "", req.region, req.language),
        "google_search", job_id, _gsearch_rows))
    return {"job_id": job_id}


@app.get("/api/gresults/{job_id}")
async def gsearch_results(job_id: str, limit: int = 2000):
    rows = [d async for d in gresults.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/bbb")
async def bbb_start(req: BBBRequest):
    """BBB Business Scraper (bbb.org). Queries may be search terms or bbb.org URLs."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "bbb", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        bbb.run_job(job_id, queries, req.limit),
        "bbb_business", job_id, _bbb_rows))
    return {"job_id": job_id}


@app.get("/api/bbb/results/{job_id}")
async def bbb_results(job_id: str, limit: int = 2000):
    rows = [d async for d in bbbresults.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/g2")
async def g2_start(req: G2Request):
    """G2 Reviews Scraper (g2.com). Queries may be product-reviews URLs, product URLs, or slugs."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "g2", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        g2.run_job(job_id, queries, req.limit, req.sort or ""),
        "g2_reviews", job_id, _g2_rows))
    return {"job_id": job_id}


@app.get("/api/g2/results/{job_id}")
async def g2_results(job_id: str, limit: int = 4000):
    rows, _ = await _g2_rows(job_id)
    return rows[:limit]


@app.post("/api/glassdoor-jobs")
async def glassdoor_jobs_start(req: GlassdoorJobsRequest):
    """Glassdoor Job Scraper — jobs from a glassdoor.com job-search URL."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Glassdoor job-search URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "glassdoor_jobs", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        glassdoor_jobs.run_job(job_id, queries, req.limit),
        "glassdoor_jobs", job_id, _gjobs_rows))
    return {"job_id": job_id}


@app.get("/api/glassdoor-jobs/results/{job_id}")
async def glassdoor_jobs_results(job_id: str, limit: int = 4000):
    rows, _ = await _gjobs_rows(job_id)
    return rows[:limit]


@app.post("/api/glassdoor-companies")
async def glassdoor_companies_start(req: GlassdoorCompaniesRequest):
    """Glassdoor Company Search — company search results from Glassdoor (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one company name or Glassdoor URL is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "glassdoor_companies", "queries": queries, "limit": lim,
        "domain": req.domain, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        glassdoor_companies.run_job(job_id, queries, lim, req.domain),
        "glassdoor_companies", job_id, _gco_rows))
    return {"job_id": job_id}


@app.get("/api/glassdoor-companies/results/{job_id}")
async def glassdoor_companies_results(job_id: str, limit: int = 4000):
    rows, _ = await _gco_rows(job_id)
    return rows[:limit]


@app.post("/api/glassdoor-reviews")
async def glassdoor_reviews_start(req: GlassdoorReviewsRequest):
    """Glassdoor Reviews Scraper — company reviews from a glassdoor.com reviews URL."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Glassdoor reviews URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "glassdoor_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        glassdoor_reviews.run_job(job_id, queries, req.limit, req.sort or ""),
        "glassdoor_reviews", job_id, _gdreviews_rows))
    return {"job_id": job_id}


@app.get("/api/glassdoor-reviews/results/{job_id}")
async def glassdoor_reviews_results(job_id: str, limit: int = 4000):
    rows, _ = await _gdreviews_rows(job_id)
    return rows[:limit]


@app.post("/api/walmart")
async def walmart_start(req: WalmartProductsRequest):
    """Walmart Products Scraper — products from walmart.com /ip/ URLs."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Walmart product URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "walmart", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        walmart.run_job(job_id, queries, req.limit),
        "walmart_products", job_id, _walmart_rows))
    return {"job_id": job_id}


@app.get("/api/walmart/results/{job_id}")
async def walmart_results(job_id: str, limit: int = 4000):
    rows, _ = await _walmart_rows(job_id)
    return rows[:limit]


@app.post("/api/walmart-reviews")
async def walmart_reviews_start(req: WalmartReviewsRequest):
    """Walmart Reviews Scraper — reviews from walmart.com product URLs."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Walmart product URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "walmart_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        walmart_rv.run_job(job_id, queries, req.limit, req.sort or ""),
        "walmart_reviews", job_id, _walmartrv_rows))
    return {"job_id": job_id}


@app.get("/api/walmart-reviews/results/{job_id}")
async def walmart_reviews_results(job_id: str, limit: int = 4000):
    rows, _ = await _walmartrv_rows(job_id)
    return rows[:limit]


@app.post("/api/youtube-channels")
async def youtube_channels_start(req: YouTubeChannelsRequest):
    """YouTube Channels Scraper — channel details from URLs or handles."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one channel URL or handle is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "youtube_channels", "queries": queries,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        yt_channels.run_job(job_id, queries),
        "youtube_channels", job_id, _ytchannels_rows))
    return {"job_id": job_id}


@app.get("/api/youtube-channels/results/{job_id}")
async def youtube_channels_results(job_id: str, limit: int = 4000):
    rows, _ = await _ytchannels_rows(job_id)
    return rows[:limit]


@app.post("/api/airbnb-reviews")
async def airbnb_reviews_start(req: AirbnbReviewsRequest):
    """Airbnb Reviews Scraper — reviews from airbnb.com room URLs or listing ids."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Airbnb room URL or listing id is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "airbnb_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        airbnb_rv.run_job(job_id, queries, req.limit, req.sort or ""),
        "airbnb_reviews", job_id, _airbnb_rows))
    return {"job_id": job_id}


@app.get("/api/airbnb-reviews/results/{job_id}")
async def airbnb_reviews_results(job_id: str, limit: int = 4000):
    rows, _ = await _airbnb_rows(job_id)
    return rows[:limit]


@app.post("/api/airbnb-search")
async def airbnb_search_start(req: AirbnbSearchRequest):
    """Airbnb Search Scraper — listings from an airbnb.com search URL or location (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Airbnb search URL or location is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "airbnb_search", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(airbnb.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/airbnb-search/results/{job_id}")
async def airbnb_search_results_get(job_id: str, limit: int = 2000):
    rows = [d async for d in airbnb_search_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps")
async def gmaps_start(req: GoogleMapsRequest):
    """Google Maps Data Scraper — places from Google's official Places API (New). API-only.
    Builds "<category> in <location>" queries from the categories × locations cross-product."""
    cats = [c.strip() for c in req.categories if c and c.strip()]
    locs = [l.strip() for l in req.locations if l and l.strip()]
    if not cats:
        raise HTTPException(400, "at least one category/brand is required")
    queries = [f"{c} in {l}" for c in cats for l in locs] if locs else cats
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps", "queries": queries, "categories": cats,
        "locations": locs, "limit": req.limit, "region": req.region, "language": req.language,
        "filters": req.filters, "skip": req.skip, "dedupe": req.dedupe,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(gmaps.run_job(job_id, queries, req.limit, req.region, req.language,
                                      req.filters, req.skip, req.dedupe))
    return {"job_id": job_id}


@app.get("/api/gmaps/results/{job_id}")
async def gmaps_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gmaps_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-domains")
async def gmaps_domains_start(req: GoogleMapsDomainsRequest):
    """Google Maps Search by Domains — find the Google Maps place that owns each domain/URL."""
    domains = [d.strip() for d in req.domains if d and d.strip()]
    if not domains:
        raise HTTPException(400, "at least one domain or URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_domains", "domains": domains, "limit": req.limit,
        "region": req.region, "language": req.language,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(gmaps.run_job_domains(job_id, domains, req.limit, req.region, req.language))
    return {"job_id": job_id}


@app.get("/api/gmaps-domains/results/{job_id}")
async def gmaps_domains_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gmaps_domain_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/bbb-reviews")
async def bbb_reviews_start(req: BBBReviewsRequest):
    """BBB Business Reviews Scraper — customer reviews from a bbb.org business URL."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one bbb.org reviews/profile URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "bbb_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        bbb_reviews.run_job(job_id, queries, req.limit, req.sort),
        "bbb_reviews", job_id, _bbbreviews_rows))
    return {"job_id": job_id}


@app.get("/api/bbb-reviews/results/{job_id}")
async def bbb_reviews_results(job_id: str, limit: int = 4000):
    rows = [d async for d in bbbreviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/expedia")
async def expedia_start(req: ExpediaRequest):
    """Expedia Search Scraper — hotels from an expedia.com Hotel-Search URL (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Expedia Hotel-Search URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "expedia", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(expedia.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/expedia/results/{job_id}")
async def expedia_results_get(job_id: str, limit: int = 2000):
    rows = [d async for d in expedia_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/trustpilot")
async def trustpilot_start(req: TrustpilotRequest):
    """Trustpilot Scraper — companies from a trustpilot.com URL or company ID (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Trustpilot URL or company ID is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "trustpilot", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(trustpilot.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/trustpilot/results/{job_id}")
async def trustpilot_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in trustpilot_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/trustpilot-search")
async def trustpilot_search_start(req: TrustpilotSearchRequest):
    """Trustpilot Search Scraper — companies matching a keyword search (browser-rendered)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one search keyword is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "trustpilot_search", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(trustpilot_search.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/trustpilot-search/results/{job_id}")
async def trustpilot_search_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in trustpilot_search_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/trustpilot-reviews")
async def trustpilot_reviews_start(req: TrustpilotReviewsRequest):
    """Trustpilot Reviews Summary — reviews from a trustpilot.com /review/ page (browser, proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Trustpilot /review/ URL or company id is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "trustpilot_reviews", "queries": queries, "limit": req.limit,
        "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(trustpilot_reviews_mod.run_job(job_id, queries, req.limit, req.language))
    return {"job_id": job_id}


@app.get("/api/trustpilot-reviews/results/{job_id}")
async def trustpilot_reviews_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in trustpilot_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/trustpilot-monitoring")
async def trustpilot_monitoring_start(req: TrustpilotMonitorRequest):
    """Trustpilot Reviews Monitoring — create a recurring monitor + run the first scan now."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one Trustpilot /review/ URL or company id is required")
    if req.frequency not in monitor.FREQ_DAYS:
        raise HTTPException(400, f"frequency must be one of {list(monitor.FREQ_DAYS)}")
    res = await monitor.start_monitor(queries, req.frequency, req.email, req.threshold,
                                      req.language, req.limit)
    res["frequency"] = monitor.FREQ_LABEL[req.frequency]
    return res


@app.get("/api/trustpilot-monitoring/{monitor_id}")
async def trustpilot_monitoring_get(monitor_id: str):
    doc = await monitors.find_one({"monitor_id": monitor_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "monitor not found")
    return doc


@app.post("/api/gmaps-monitoring")
async def gmaps_monitoring_start(req: GMapsMonitorRequest):
    """Google Maps Reviews Monitoring — create a recurring monitor + run the first scan now.
    Reuses the Google Maps Reviews scraper (official Places API, max 5 reviews/place)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one place_id, Maps URL, or 'category, city' query is required")
    if req.frequency not in monitor.FREQ_DAYS:
        raise HTTPException(400, f"frequency must be one of {list(monitor.FREQ_DAYS)}")
    res = await monitor.start_monitor(queries, req.frequency, req.email, req.threshold,
                                      req.language, req.limit, kind="gmaps", sort=req.sort)
    res["frequency"] = monitor.FREQ_LABEL[req.frequency]
    return res


@app.get("/api/gmaps-monitoring/{monitor_id}")
async def gmaps_monitoring_get(monitor_id: str):
    doc = await monitors.find_one({"monitor_id": monitor_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "monitor not found")
    return doc


@app.post("/api/hotels")
async def hotels_start(req: HotelsRequest):
    """Hotels Search Scraper — hotels from a hotels.com Hotel-Search URL (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one hotels.com Hotel-Search URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "hotels", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(hotels.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/hotels/results/{job_id}")
async def hotels_results_get(job_id: str, limit: int = 2000):
    rows = [d async for d in hotels_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/hotels-reviews")
async def hotels_reviews_start(req: HotelsReviewsRequest):
    """Hotels Reviews Scraper — guest reviews from a hotels.com hotel URL (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one hotels.com hotel URL is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "hotels_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(hotels_reviews_mod.run_job(job_id, queries, req.limit, req.sort))
    return {"job_id": job_id}


@app.get("/api/hotels-reviews/results/{job_id}")
async def hotels_reviews_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in hotels_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/expedia-reviews")
async def expedia_reviews_start(req: ExpediaReviewsRequest):
    """Expedia Reviews Scraper — guest reviews from an expedia.com hotel URL (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one expedia.com hotel URL or hotel id is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "expedia_reviews", "queries": queries, "limit": req.limit,
        "sort": req.sort, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(expedia_reviews_mod.run_job(job_id, queries, req.limit, req.sort))
    return {"job_id": job_id}


@app.get("/api/expedia-reviews/results/{job_id}")
async def expedia_reviews_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in expedia_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-reviews")
async def gmaps_reviews_start(req: GMapsReviewsRequest):
    """Google Maps Reviews Scraper — reviews from Google Maps places (internal RPC, proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one place query / URL / id is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_reviews", "queries": queries, "sort": req.sort,
        "limit": req.limit, "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    lim = None if (req.limit or 0) == 0 else req.limit
    asyncio.create_task(gmaps_reviews_mod.run_job(job_id, queries, req.sort, lim, req.language))
    return {"job_id": job_id}


@app.get("/api/gmaps-reviews/results/{job_id}")
async def gmaps_reviews_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.get("/api/gmaps-categories")
async def gmaps_categories():
    """Google Maps category list (from app/categories.xlsx) for the Categories/brands dropdown."""
    return gmaps_reviews_mod.categories()


@app.post("/api/gnews")
async def gnews_start(req: GNewsRequest):
    """Google Search News Scraper — news articles for a query via Google News RSS (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    lim = None if (req.limit or 0) == 0 else req.limit
    await jobs.insert_one({
        "job_id": job_id, "kind": "gnews", "queries": queries, "limit": lim,
        "country": req.country, "date_range": req.date_range, "language": req.language,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gnews.run_job(job_id, queries, lim, req.date_range or "", req.country, req.language),
        "google_news", job_id, _gnews_rows))
    return {"job_id": job_id}


@app.get("/api/gnews-results/{job_id}")
async def gnews_results_get(job_id: str, limit: int = 2000):
    rows = [d async for d in gnews_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gimages")
async def gimages_start(req: GImagesRequest):
    """Google Search Images Scraper — image results via DuckDuckGo's image JSON API (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gimages", "queries": queries, "limit": lim,
        "country": req.country, "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gimages.run_job(job_id, queries, lim, req.country, req.language),
        "google_images", job_id, _gimages_rows))
    return {"job_id": job_id}


@app.get("/api/gimages-results/{job_id}")
async def gimages_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gimages_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gvideos")
async def gvideos_start(req: GVideosRequest):
    """Google Search Videos Scraper — video results via Bing video search (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gvideos", "queries": queries, "limit": lim,
        "country": req.country, "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gvideos.run_job(job_id, queries, lim, req.country, req.language),
        "google_videos", job_id, _gvideos_rows))
    return {"job_id": job_id}


@app.get("/api/gvideos-results/{job_id}")
async def gvideos_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gvideos_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gevents")
async def gevents_start(req: GEventsRequest):
    """Google Search Events Scraper — event listings via Google's events pack (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    pages = req.limit or 1
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gevents", "queries": queries, "limit": pages,
        "country": req.country, "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gevents.run_job(job_id, queries, pages, req.country, req.language),
        "google_events", job_id, _gevents_rows))
    return {"job_id": job_id}


@app.get("/api/gevents-results/{job_id}")
async def gevents_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gevents_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gcareers")
async def gcareers_start(req: GCareersRequest):
    """Google Search Careers Scraper — jobs from a careers.google.com search (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gcareers", "queries": queries, "limit": lim,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gcareers.run_job(job_id, queries, lim),
        "google_careers", job_id, _gcareers_rows))
    return {"job_id": job_id}


@app.get("/api/gcareers-results/{job_id}")
async def gcareers_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gcareers_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gtrends")
async def gtrends_start(req: GTrendsRequest):
    """Google Trends Scraper — interest-by-region via Google Trends' free internal API (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gtrends", "queries": queries, "geo": req.geo,
        "timeframe": req.timeframe, "resolution": req.resolution,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gtrends.run_job(job_id, queries, req.geo, req.timeframe, req.resolution),
        "google_trends", job_id, _gtrends_rows))
    return {"job_id": job_id}


@app.get("/api/gtrends-results/{job_id}")
async def gtrends_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gtrends_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/linkedin-companies")
async def linkedin_companies_start(req: LinkedInCompaniesRequest):
    """LinkedIn Companies Scraper — company details from linkedin.com/company pages (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "linkedin_companies", "queries": queries,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        linkedin_companies.run_job(job_id, queries),
        "linkedin_companies", job_id, _linkedin_companies_rows))
    return {"job_id": job_id}


@app.get("/api/linkedin-companies-results/{job_id}")
async def linkedin_companies_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in linkedin_companies_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/linkedin-posts")
async def linkedin_posts_start(req: LinkedInPostsRequest):
    """LinkedIn Posts Scraper — recent posts from a company profile (auth + proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "linkedin_posts", "queries": queries, "limit": lim,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        linkedin_posts.run_job(job_id, queries, lim),
        "linkedin_posts", job_id, _linkedin_posts_rows))
    return {"job_id": job_id}


@app.get("/api/linkedin-posts-results/{job_id}")
async def linkedin_posts_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in linkedin_posts_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gsjobs")
async def gsjobs_start(req: GSJobsRequest):
    """Google Search Jobs Scraper — job listings via Google Jobs (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gsjobs", "queries": queries, "pages": req.pages,
        "language": req.language, "region": req.region, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gsjobs.run_job(job_id, queries, req.pages, req.language, req.region),
        "google_jobs", job_id, _gsjobs_rows))
    return {"job_id": job_id}


@app.get("/api/gsjobs-results/{job_id}")
async def gsjobs_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gsjobs_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gshop")
async def gshop_start(req: GShopRequest):
    """Google Search Shopping Scraper — product results via Google Shopping (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gshop", "queries": queries, "limit": lim,
        "language": req.language, "region": req.region, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gshop.run_job(job_id, queries, lim, req.language, req.region),
        "google_shopping", job_id, _gshop_rows))
    return {"job_id": job_id}


@app.get("/api/gshop-results/{job_id}")
async def gshop_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in gshop_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gsreviews")
async def gsreviews_start(req: GShopReviewsRequest):
    """Google Shopping Reviews Scraper — reviews for a list of Google Shopping products (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one product link or product id is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gsreviews", "queries": queries, "limit": lim,
        "language": req.language, "region": req.region, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gsreviews.run_job(job_id, queries, lim, req.language, req.region),
        "google_shopping_reviews", job_id, _gsr_rows))
    return {"job_id": job_id}


@app.get("/api/gsreviews-results/{job_id}")
async def gsreviews_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gsreviews_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/linkedin-profiles")
async def linkedin_profiles_start(req: LinkedInProfilesRequest):
    """LinkedIn Profiles Scraper — public-profile data for a list of profile URLs/ids (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one profile URL or id is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "linkedin_profiles", "queries": queries, "status": "running",
        "total_scraped": 0, "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        linkedin_profiles_mod.run_job(job_id, queries),
        "linkedin_profiles", job_id, _lip_rows))
    return {"job_id": job_id}


@app.get("/api/linkedin-profiles-results/{job_id}")
async def linkedin_profiles_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in linkedin_profiles.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gflights")
async def gflights_start(req: GFlightsRequest):
    """Google Search Flights Scraper — flight results from Google Flights (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one origin,destination pair is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gflights", "queries": queries, "limit": lim,
        "departure_date": req.departure_date, "return_date": req.return_date,
        "language": req.language, "region": req.region, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gflights.run_job(job_id, queries, req.departure_date, req.return_date, lim,
                         req.language, req.region),
        "google_flights", job_id, _gfl_rows))
    return {"job_id": job_id}


@app.get("/api/gflights-results/{job_id}")
async def gflights_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gflights_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-autocomplete")
async def gmaps_autocomplete_start(req: GMapsAutocompleteRequest):
    """Google Maps Autocomplete — suggestion lists for Maps search queries (works on the free pool)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_autocomplete", "queries": queries,
        "coordinates": req.coordinates, "language": req.language, "region": req.region,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_autocomplete_mod.run_job(job_id, queries, req.coordinates, req.language, req.region),
        "gmaps_autocomplete", job_id, _gma_rows))
    return {"job_id": job_id}


@app.get("/api/gmaps-autocomplete-results/{job_id}")
async def gmaps_autocomplete_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_autocomplete.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gsearch-autocomplete")
async def gsearch_autocomplete_start(req: GSearchAutocompleteRequest):
    """Google Search Autocomplete — suggestion lists for search queries (works on the free pool)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one query is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gsearch_autocomplete", "queries": queries,
        "language": req.language, "region": req.region, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_autocomplete_mod.run_job_search(job_id, queries, req.language, req.region),
        "gsearch_autocomplete", job_id, _gsa_rows))
    return {"job_id": job_id}


@app.get("/api/gsearch-autocomplete-results/{job_id}")
async def gsearch_autocomplete_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gsearch_autocomplete.find(
        {"job_id": job_id}, {"_id": 0, "job_id": 0, "coordinates": 0})]
    return rows[:limit]


@app.post("/api/gplay")
async def gplay_start(req: GPlayRequest):
    """Google Play Reviews Scraper — app reviews via Play's internal batchexecute API (proxy)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one app id or Play Store URL is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gplay", "queries": queries, "limit": lim, "sort": req.sort,
        "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gplay.run_job(job_id, queries, lim, req.sort, req.language),
        "google_play_reviews", job_id, _gplay_rows))
    return {"job_id": job_id}


@app.get("/api/gplay-results/{job_id}")
async def gplay_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gplay_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gplay-monitoring")
async def gplay_monitoring_start(req: GPlayMonitorRequest):
    """Google Play Reviews Monitoring — recurring app-review scan + email report. Reuses the Play
    Reviews scraper (internal batchexecute API, proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one app id or Play Store URL is required")
    if req.frequency not in monitor.FREQ_DAYS:
        raise HTTPException(400, f"frequency must be one of {list(monitor.FREQ_DAYS)}")
    res = await monitor.start_monitor(queries, req.frequency, req.email, req.threshold,
                                      req.language, req.limit, kind="gplay", sort=req.sort)
    res["frequency"] = monitor.FREQ_LABEL[req.frequency]
    return res


@app.get("/api/gplay-monitoring/{monitor_id}")
async def gplay_monitoring_get(monitor_id: str):
    doc = await monitors.find_one({"monitor_id": monitor_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "monitor not found")
    return doc


@app.post("/api/gmaps-traffic")
async def gmaps_traffic_start(req: GMapsTrafficRequest):
    """Google Maps Traffic Scraper — directions (travel time + distance) between two points (proxy-only)."""
    pairs = [{"start": (p.start or "").strip(), "stop": (p.stop or "").strip()}
             for p in req.pairs if (p.start or "").strip() and (p.stop or "").strip()]
    if not pairs:
        raise HTTPException(400, "at least one Start + Stop location pair is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_traffic", "pairs": pairs, "travel_mode": req.travel_mode,
        "time_from": req.time_from, "time_to": req.time_to, "interval_min": req.interval_min,
        "status": "running", "total_scraped": 0, "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_traffic_mod.run_job(job_id, pairs, req.time_from, req.time_to, req.interval_min,
                                  req.travel_mode),
        "google_maps_traffic", job_id, _gmaps_traffic_rows))
    return {"job_id": job_id}


@app.get("/api/gmaps-traffic-results/{job_id}")
async def gmaps_traffic_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_traffic.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-directory")
async def gmaps_directory_start(req: GMapsDirectoryRequest):
    """Google Maps Directory Places — business listings from a search/place (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one 'category, city' query, Maps URL, or place_id is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_directory", "queries": queries, "limit": lim,
        "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_directory_mod.run_job(job_id, queries, lim, req.language),
        "google_maps_directory", job_id, _gmaps_directory_rows))
    return {"job_id": job_id}


@app.get("/api/gmaps-directory-results/{job_id}")
async def gmaps_directory_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_directory.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-contrib")
async def gmaps_contrib_start(req: GMapsContribRequest):
    """Google Maps Contributor Reviews Scraper — all reviews a contributor left (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one contributor ID or /contrib/<id> URL is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_contrib", "queries": queries, "limit": lim,
        "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_contrib.run_job(job_id, queries, lim, req.language),
        "google_maps_contrib", job_id, _gmaps_contrib_rows))
    return {"job_id": job_id}


@app.get("/api/gmaps-contrib-results/{job_id}")
async def gmaps_contrib_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_contrib_reviews.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/gmaps-photos")
async def gmaps_photos_start(req: GMapsPhotosRequest):
    """Google Maps Photos Scraper — all photos from a place (proxy-only, free DOM scrape)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one place_id, Maps URL, or 'category, city' query is required")
    lim = None if (req.limit or 0) == 0 else req.limit
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "gmaps_photos", "queries": queries, "limit": lim,
        "language": req.language, "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(_run_and_archive(
        gmaps_photos_mod.run_job(job_id, queries, lim, req.language),
        "google_maps_photos", job_id, _gmaps_photos_rows))
    return {"job_id": job_id}


@app.get("/api/gmaps-photos-results/{job_id}")
async def gmaps_photos_results_get(job_id: str, limit: int = 5000):
    rows = [d async for d in gmaps_photos.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]


@app.post("/api/homedepot")
async def homedepot_start(req: HomeDepotRequest):
    """Home Depot Products Scraper — product listings from a homedepot.com URL (proxy-only)."""
    queries = [q.strip() for q in req.queries if q and q.strip()]
    if not queries:
        raise HTTPException(400, "at least one homedepot.com URL or keyword is required")
    job_id = uuid.uuid4().hex
    await jobs.insert_one({
        "job_id": job_id, "kind": "homedepot", "queries": queries, "limit": req.limit,
        "status": "running", "total_scraped": 0,
        "started_at": datetime.utcnow(), "finished_at": None,
    })
    asyncio.create_task(homedepot.run_job(job_id, queries, req.limit))
    return {"job_id": job_id}


@app.get("/api/homedepot/results/{job_id}")
async def homedepot_results_get(job_id: str, limit: int = 3000):
    rows = [d async for d in homedepot_results.find({"job_id": job_id}, {"_id": 0, "job_id": 0})]
    return rows[:limit]
@app.post("/api/stop/{job_id}")
async def stop(job_id: str):
    doc = await jobs.find_one({"job_id": job_id}, {"_id": 0, "status": 1})
    if not doc:
        raise HTTPException(404, "job not found")
    if doc.get("status") == "running":
        request_stop(job_id)
    return {"job_id": job_id, "stopping": True}


@app.get("/api/status/{job_id}")
async def status(job_id: str):
    doc = await jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "job not found")
    return doc


@app.get("/api/results/{job_id}")
async def results(job_id: str, limit: int = 200):
    job = await jobs.find_one({"job_id": job_id}, {"_id": 0, "sort": 1, "categories": 1}) or {}
    # fetch all, then apply the job's Category filter + Sort, then cap (sort/filter need the
    # full set before slicing — a DB-level .limit() would truncate before the view applies)
    rows = [d async for d in businesses.find({"job_id": job_id}, {"_id": 0})]
    rows = apply_view(rows, job.get("sort"), job.get("categories"))
    return rows[:limit]


@app.get("/api/export/{job_id}")
async def export(job_id: str):
    doc = await jobs.find_one({"job_id": job_id})
    if not doc or not doc.get("export_path"):
        raise HTTPException(404, "export not ready")
    import os
    return FileResponse(
        doc["export_path"],
        media_type="application/json",
        filename=os.path.basename(doc["export_path"]),
    )


# Catch-all (declared LAST so it never shadows /api/* or other routes): serve the SPA so a direct
# load / refresh of a per-service URL like /Emails-Contacts-Scraper works (client-side routing).
@app.get("/{full_path:path}", response_class=HTMLResponse)
async def spa(full_path: str):
    with open("static/index.html", encoding="utf-8") as f:
        return f.read()
